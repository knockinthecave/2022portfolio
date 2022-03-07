from openpyxl import load_workbook
from tkinter import *
import schedule

wb = load_workbook("data1.xlsx", data_only=True)
ws = wb.active

name = []

for cell in ws['A']:
    if cell.value == '거래처':
        
        continue
    if cell.value is None:
        break

    name.append(cell.value)

cycle = []

for cell in ws['C']:
    if cell.value == '배차주기':
        continue
    if cell.value is None:
        break

    cycle.append(cell.value)

wb = load_workbook("data2.xlsx", data_only=True)
ws = wb.active

name1 = []

for cell in ws['C']:
    if cell.value == '거래처명':
        continue
    if cell.value is None:
        break
    name1.append(cell.value)

distance = []

for cell in ws['D']:
    if cell.value == '거리[km]':
        continue
    if cell.value is None:
        break
    distance.append(cell.value)

open_time = []

for cell in ws['B']:
    if cell.value == '오픈':
        continue
    if cell.value is None:
        break
    open_time.append(cell.value)

month_dict = dict(zip(name, cycle))
open_time_dict = dict(zip(name1, open_time))
dis_dict = dict(zip(name1, distance))

m_name_list = []
m_cycle_list = []
m_ot_list = []
m_dis_list = []

for month_name in name:
    if month_name in name1:
        if month_dict[month_name] == '?':
            month_dict[month_name] = 1000
            m_cycle_list.append(month_dict[month_name])
        else:
            m_cycle_list.append(month_dict[month_name])
            
        m_name_list.append(month_name)
        
        m_ot_list.append(open_time_dict[month_name])
        m_dis_list.append(dis_dict[month_name])

def info(i):
    root = Tk()
    root.title("배차알리미")
    root.geometry("400x150")
    root.resizable(False, False)
    label1 = Label(root, text="거래처명 = " + str(m_name_list[i]))
    label2 = Label(root, text="배차간격 = " + str(m_cycle_list[i]) + "일")
    label3 = Label(root, text="운행거리 = " + str(m_dis_list[i]) + "km")
    label4 = Label(root, text="오픈시간 = " + str(m_ot_list[i]) + "시")

    label1.pack()
    label2.pack()
    label3.pack()
    label4.pack()

    root.mainloop()

for a in range(len(m_cycle_list)):
    schedule.every(m_cycle_list[a]).days.do(info,a)
    schedule.run_pending()
